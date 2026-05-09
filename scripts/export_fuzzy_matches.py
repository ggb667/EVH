import csv
import difflib
import re

from openpyxl import load_workbook


XLSX = "EVHInventorySuppliers.xlsx"
STOCK = "Stockroom · Instinct Stockroom.csv"
OUT = "docs/evh_fuzzy_matches_xls_left_stockroom_right.csv"


def norm(value: str) -> str:
    value = (value or "").lower().strip()
    value = value.lstrip("📦 ")
    value = value.replace("&", " and ")
    return re.sub(r"[^a-z0-9]+", "", value)


def main() -> None:
    with open(STOCK, newline="", encoding="utf-8-sig") as f:
        stock_rows = list(csv.DictReader(f))

    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb.active
    wb_rows: list[tuple[str, str, str, str]] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx == 1:
            continue
        code = "" if len(row) < 1 or row[0] is None else str(row[0]).strip()
        desc = "" if len(row) < 2 or row[1] is None else str(row[1]).strip()
        supplier = "" if len(row) < 3 or row[2] is None else str(row[2]).strip()
        mfg = "" if len(row) < 4 or row[3] is None else str(row[3]).strip()
        if code or desc:
            wb_rows.append((code, desc, supplier, mfg))

    results = []
    for srow in stock_rows:
        sid = srow["PIMS ID"].strip()
        sprod = srow["Product"].strip()
        sn = norm(sprod)
        best = None
        best_score = 0.0
        for code, desc, supplier, mfg in wb_rows:
            score = difflib.SequenceMatcher(None, sn, norm(desc)).ratio()
            if score > best_score:
                best_score = score
                best = (code, desc, supplier, mfg)
        if best and best_score >= 0.84:
            code, desc, supplier, mfg = best
            results.append(
                {
                    "Workbook ID": code,
                    "Workbook Product": desc,
                    "Workbook Supplier": supplier,
                    "Stockroom ID": sid,
                    "Stockroom Product": sprod,
                    "Score": f"{best_score:.3f}",
                }
            )

    with open(OUT, "w", newline="", encoding="utf-8") as f:
        fieldnames = [
            "Workbook ID",
            "Workbook Product",
            "Workbook Supplier",
            "Stockroom ID",
            "Stockroom Product",
            "Score",
        ]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(results)

    print(OUT)
    print(f"rows={len(results)}")
    for row in results[:25]:
        print(
            f"{row['Workbook ID']} | {row['Workbook Product']} | {row['Workbook Supplier']} || "
            f"{row['Stockroom ID']} | {row['Stockroom Product']} | {row['Score']}"
        )


if __name__ == "__main__":
    main()
