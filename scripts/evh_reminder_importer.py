#!/usr/bin/env python3
"""
EVH reminder importer

Purpose:
1. Read the grouped reminder spreadsheet layout.
2. Parse one patient group into one patient identity record plus many reminder rows.
3. Apply source-to-Instinct mappings.
4. Exclude rows that should not create reminders.
5. Support dry-run inspection and JSON export.
6. Provide adapter hooks for the already-proven EVH auth / patient lookup / patient patch flow.

Important:
- This file intentionally does NOT guess the final Instinct reminder endpoint or payload shape.
- Wire the InstinctApiAdapter methods to the existing EVH implementation that already:
  a. authenticates
  b. obtains token
  c. finds patient
  d. PATCHes patient data

Expected spreadsheet columns:
A Client
B Name
C Phone No.
D Patient
E Species
F Breed
G Code
H Treatments/Items/Diagnoses Due
I Due Date

Grouped layout:
- First row of a group contains patient identity fields AND the first reminder row.
- Following rows in the same group have A-F blank and G-I populated.
- A blank row ends the group.
"""

from __future__ import annotations

import argparse
import dataclasses
import json
import logging
import os
import re
from collections.abc import Iterable
from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path
from typing import Any, Optional

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency for spreadsheet mode only
    load_workbook = None


LOGGER = logging.getLogger("evh_reminder_importer")


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class SourceReminderRow:
    code: str
    source_label: str
    due_date_raw: Any
    row_number: int


@dataclass
class ParsedPatientGroup:
    client: str
    client_name: str
    phone_no: str
    patient_name: str
    species: str
    breed: str
    header_row_number: int
    reminders: list[SourceReminderRow] = field(default_factory=list)


@dataclass
class ReminderCandidate:
    source_code: str
    source_label: str
    mapped_label: str
    due_date: str
    row_number: int


@dataclass
class SkippedRow:
    source_code: str
    source_label: str
    reason: str
    row_number: int


@dataclass
class PatientImportPlan:
    patient: ParsedPatientGroup
    valid_reminders: list[ReminderCandidate] = field(default_factory=list)
    skipped_rows: list[SkippedRow] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Mapping and exclusions
# ---------------------------------------------------------------------------

# Finalized many-to-one source mapping based on the conversation.
SOURCE_TO_INSTINCT_MAP: dict[str, str] = {
    "Annual Wellness Exam": "Annual Exam",
    "Bloodwork - annual w/ Heartworm": "Annual Bloodwork",
    "Bordetella Oral Vaccine": "Bordetella Oral Parainfluenza Vaccine",
    "Credelio for Cats 4.1 - 17.0#": "Flea / Tick Prevention",
    "DA2P Annual (w/o Lepto)": "DA2P Annual ( w/o Lepto)",
    "DA2PP + Leptospirosis 4 Annual": "DA2P + Leptospirosis 4",
    "Dhlp Parvo - Corona K4": "DHLP Vaccine",
    "Feline Leukemia Annual": "Feline Leukemia (FeLV) Vaccine",
    "Feline Rabies 1 yr/Purevax": "Rabies Vaccine (Feline)",
    "Fvrcpc Annual": "FVRCP Vaccine",
    "General Senior Profile (IRL 78": "Senior Blood Work",
    "Heartworm Test - Previous Vete": "Heartworm Test",
    "In House Flex 4 Rapid Test": "Heartworm Test",
    "In House Imagyst Fecal/Oocysts": "Fecal Analysis",
    "Inject - Cytopoint": "Cytopoint Injection",
    "Inject - ProHeart 12": "ProHeart 12 Injection",
    "Inject - Solensia": "Solensia Injection",
    "Leptospirosis 4 Vaccine Annual": "Leptospirosis Vaccine",
    "Lyme Vaccine Annual": "Borrelia burgdorferi (Lyme) Vaccine",
    "Nexgard COMBO for Cats 5.6 - 16#": "Flea / Tick / Heartworm Prevention",
    "Nexgard Plus for Dogs 17 - 33#": "Flea / Tick / Heartworm Prevention",
    "Physical Exam": "Annual Exam",
    "Rabies Canine 1 yr": "Rabies Vaccine (Canine)",
    "Rabies Canine 3 yr": "Rabies 3 yr Vaccine (Canine)",
    "Rabies Canine 3yr": "Rabies 3 yr Vaccine (Canine)",
    "Revolution Plus Cat 11.1 - 22#": "Flea / Tick / Heartworm Prevention",
    "Revolution Plus Cat 5.6 - 11#": "Flea / Tick / Heartworm Prevention",
    "Simparica Trio 44.1 - 88#": "Flea / Tick / Heartworm Prevention",
    "Thyroid T4 only": "Thyroid Level",
    "Antech Canine Urine Microalbum": "Urinalysis",
    "Antech Feline Urine Microalbum": "Urinalysis",
    "Urinalysis (In House)": "Urinalysis",
    "Antech Sr Canine Wellness w/ A": "Senior Blood Work",
    "Antech Thyroid (Total T4)": "Thyroid Level",
    "Antech Thyroid T4, FT4, cTSH": "Thyroid Level",
    "Antech Free T4 (Equilibrium Di": "Thyroid Level",
    "Antech Phenobarb Levels ( Add-": "Phenobarbital Level",
    "Antech CBC, CHEM25, T4, freeT4": "Comprehensive Bloodwork",
    "Antech Urine Protein / Creatin": "Urinalysis",
    "In House Abaxis T4 / Cholester": "Thyroid Level",
    "Antech Feline Comp Well Screen": "Annual Bloodwork",
    "Antech Cholesterol/Triglycerid": "Annual Bloodwork",
    "Antech Accuplex 4 ( HW plus Ti": "Heartworm Test",
    "Antech Canine Heartworm ANTIGE": "Heartworm Test",
    "Dermatonin Implant": "Dermatonin Implant",
    "Dermatonin Implant 8.0mg": "Dermatonin Implant",
}


EXCLUDED_SOURCE_VALUES: set[str] = {
    "Antech Urine MIC Culture",
    "In House Abx CBC",
    "In House Abx Chem 17",
    "In House Glucose Curve",
    "Inject - Adequan Canine",
    "Inject - Adequan Feline",
    "Inject - B12",
    "Inject - Convenia",
    "Inject - Dexamethasone SP",
    "Inject - Depo Medrol",
    "Metacam Injection",
    "Microchip Implant",
}


# Pattern-based mappings for expanded weight-range families that should all map the same way.
PATTERN_MAPPINGS: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"^Heartgard Plus\b", re.IGNORECASE), "Heartworm Prevention"),
    (re.compile(r"^Nexgard Plus for Dogs\b", re.IGNORECASE), "Flea / Tick / Heartworm Prevention"),
    (re.compile(r"^Nexgard COMBO for Cats\b", re.IGNORECASE), "Flea / Tick / Heartworm Prevention"),
    (re.compile(r"^Revolution Plus Cat\b", re.IGNORECASE), "Flea / Tick / Heartworm Prevention"),
    (re.compile(r"^Simparica Trio\b", re.IGNORECASE), "Flea / Tick / Heartworm Prevention"),
    (re.compile(r"^Credelio for Cats\b", re.IGNORECASE), "Flea / Tick Prevention"),
    (re.compile(r"^Antech Fecal\b", re.IGNORECASE), "Fecal Analysis"),
    (re.compile(r"^Inject - ProHeart 12\b", re.IGNORECASE), "ProHeart 12 Injection"),
]


# ---------------------------------------------------------------------------
# Utility functions
# ---------------------------------------------------------------------------


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    # Collapse repeated whitespace but do not alter semantic punctuation.
    text = re.sub(r"\s+", " ", text)
    return text


def _extract_collection(payload: Any, keys: tuple[str, ...]) -> list[Any]:
    if isinstance(payload, list):
        return payload
    if not isinstance(payload, dict):
        return []

    for key in keys:
        value = payload.get(key)
        if isinstance(value, list):
            return value

    return []


def _coerce_int(value: Any) -> Optional[int]:
    if isinstance(value, int):
        return value
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    return None


def _extract_patient_id_from_reminder(reminder: Any) -> Optional[int]:
    if not isinstance(reminder, dict):
        return None

    for key in ("patientId", "patient_id"):
        patient_id = _coerce_int(reminder.get(key))
        if patient_id is not None:
            return patient_id

    patient = reminder.get("patient")
    if isinstance(patient, dict):
        return _coerce_int(patient.get("id"))

    return None


def _extract_phone_no(communication_details: Iterable[Any]) -> Optional[str]:
    for item in communication_details:
        if not isinstance(item, dict):
            continue

        value = item.get("value") or item.get("number") or item.get("phoneNumber")
        kind = (item.get("type") or item.get("kind") or "").lower()

        if value and ("phone" in kind or "mobile" in kind or kind == ""):
            return value

    return None


def _normalize_lookup_text(value: Any) -> str:
    return normalize_text(value).casefold()


def _normalize_phone_no(value: Any) -> str:
    return "".join(ch for ch in str(value or "") if ch.isdigit())


def _account_display_name(account: Any) -> Optional[str]:
    if not isinstance(account, dict):
        return None

    primary = account.get("primaryContact") or {}
    if not isinstance(primary, dict):
        return None

    first = normalize_text(primary.get("nameFirst"))
    middle = normalize_text(primary.get("nameMiddle"))
    last = normalize_text(primary.get("nameLast"))
    name_parts = [part for part in (first, middle, last) if part]
    if name_parts:
        return " ".join(name_parts)
    return None



def is_blank_row(values: Iterable[Any]) -> bool:
    return all(normalize_text(v) == "" for v in values)



def parse_due_date(value: Any) -> str:
    if value is None or value == "":
        raise ValueError("Missing due date")

    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = normalize_text(value)
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue

    raise ValueError(f"Unsupported due date format: {text!r}")



def map_source_label(source_label: str) -> Optional[str]:
    if source_label in EXCLUDED_SOURCE_VALUES:
        return None

    direct = SOURCE_TO_INSTINCT_MAP.get(source_label)
    if direct:
        return direct

    for pattern, mapped in PATTERN_MAPPINGS:
        if pattern.search(source_label):
            return mapped

    return None


# ---------------------------------------------------------------------------
# Spreadsheet parsing
# ---------------------------------------------------------------------------


def parse_grouped_spreadsheet(path: Path, worksheet_name: Optional[str] = None) -> list[ParsedPatientGroup]:
    if load_workbook is None:
        raise RuntimeError("Spreadsheet parsing requires openpyxl. Install it before reading .xlsx reminder files.")

    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[worksheet_name] if worksheet_name else workbook.active

    groups: list[ParsedPatientGroup] = []
    active_group: Optional[ParsedPatientGroup] = None

    # Assume row 1 is the header.
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_col=9, values_only=True), start=2):
        row_values = list(row[:9])
        a, b, c, d, e, f, g, h, i = row_values

        if is_blank_row(row_values):
            active_group = None
            continue

        patient_identity_present = any(normalize_text(v) for v in (a, b, c, d, e, f))
        reminder_data_present = any(normalize_text(v) for v in (g, h, i))

        if patient_identity_present:
            if not reminder_data_present:
                raise ValueError(f"Header row {row_idx} has patient identity but no reminder fields")

            active_group = ParsedPatientGroup(
                client=normalize_text(a),
                client_name=normalize_text(b),
                phone_no=normalize_text(c),
                patient_name=normalize_text(d),
                species=normalize_text(e),
                breed=normalize_text(f),
                header_row_number=row_idx,
                reminders=[],
            )
            groups.append(active_group)
        else:
            if active_group is None:
                raise ValueError(f"Continuation row {row_idx} encountered before a patient group header")

        if active_group is None:
            raise AssertionError("active_group unexpectedly None after header/continuation processing")

        code = normalize_text(g)
        source_label = normalize_text(h)
        due_date_raw = i

        if not code and not source_label and normalize_text(due_date_raw) == "":
            raise ValueError(f"Row {row_idx} is neither blank nor valid reminder data")

        active_group.reminders.append(
            SourceReminderRow(
                code=code,
                source_label=source_label,
                due_date_raw=due_date_raw,
                row_number=row_idx,
            )
        )

    return groups


# ---------------------------------------------------------------------------
# Planning logic
# ---------------------------------------------------------------------------


def build_import_plan(groups: list[ParsedPatientGroup]) -> list[PatientImportPlan]:
    plans: list[PatientImportPlan] = []

    for group in groups:
        plan = PatientImportPlan(patient=group)

        for src in group.reminders:
            label = src.source_label
            if not label:
                plan.skipped_rows.append(
                    SkippedRow(
                        source_code=src.code,
                        source_label=label,
                        reason="missing_source_label",
                        row_number=src.row_number,
                    )
                )
                continue

            try:
                due_date = parse_due_date(src.due_date_raw)
            except ValueError as exc:
                plan.skipped_rows.append(
                    SkippedRow(
                        source_code=src.code,
                        source_label=label,
                        reason=f"bad_due_date: {exc}",
                        row_number=src.row_number,
                    )
                )
                continue

            if label in EXCLUDED_SOURCE_VALUES:
                plan.skipped_rows.append(
                    SkippedRow(
                        source_code=src.code,
                        source_label=label,
                        reason="excluded_non_reminder_item",
                        row_number=src.row_number,
                    )
                )
                continue

            mapped = map_source_label(label)
            if mapped is None:
                plan.skipped_rows.append(
                    SkippedRow(
                        source_code=src.code,
                        source_label=label,
                        reason="no_mapping",
                        row_number=src.row_number,
                    )
                )
                continue

            plan.valid_reminders.append(
                ReminderCandidate(
                    source_code=src.code,
                    source_label=label,
                    mapped_label=mapped,
                    due_date=due_date,
                    row_number=src.row_number,
                )
            )

        plans.append(plan)

    return plans


# ---------------------------------------------------------------------------
# Instinct API adapter hooks
# ---------------------------------------------------------------------------

class InstinctApiAdapter:
    """
    Adapter...
    """

    def __init__(self, base_url: str, username: str, password: str) -> None:
        self.base_url = base_url
        self.username = username
        self.password = password
        self.token: Optional[str] = None
        self._reminder_counts_by_patient: Optional[dict[int, int]] = None
        self._has_patient_scoped_reminders = False

    def authenticate(self) -> str:
        import requests

        url = f"{self.base_url}/v1/auth/token"
        resp = requests.post(
            url,
            json={
                "grant_type": "client_credentials",
                "client_id": self.username,
                "client_secret": self.password,
            },
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()

        self.token = data["access_token"]
        return self.token

    def _get(self, path: str, params=None):
        import requests

        resp = requests.get(
            f"{self.base_url}{path}",
            headers={"Authorization": f"Bearer {self.token}"},
            params=params or {},
            timeout=30,
        )
        resp.raise_for_status()
        return resp.json()

    def iter_reminders(self):
        cursor = None

        while True:
            params = {"limit": 100}
            if cursor:
                params["pageCursor"] = cursor

            data = self._get("/v1/reminders", params)
            reminders = _extract_collection(data, ("reminders", "data", "items", "results"))
            for reminder in reminders:
                yield reminder

            cursor = data.get("nextPageCursor") if isinstance(data, dict) else None
            if not cursor:
                break

    def get_reminder_count_for_patient(self, patient_id: Any) -> Optional[int]:
        patient_id_int = _coerce_int(patient_id)
        if patient_id_int is None:
            return None

        if self._reminder_counts_by_patient is None:
            counts: dict[int, int] = {}
            has_patient_scoped_reminders = False

            for reminder in self.iter_reminders():
                reminder_patient_id = _extract_patient_id_from_reminder(reminder)
                if reminder_patient_id is None:
                    continue

                has_patient_scoped_reminders = True
                counts[reminder_patient_id] = counts.get(reminder_patient_id, 0) + 1

            self._reminder_counts_by_patient = counts
            self._has_patient_scoped_reminders = has_patient_scoped_reminders

        if not self._has_patient_scoped_reminders:
            return None

        return self._reminder_counts_by_patient.get(patient_id_int, 0)

    def iter_patients(self):
        cursor = None

        while True:
            params = {"limit": 100}
            if cursor:
                params["pageCursor"] = cursor

            data = self._get("/v1/patients", params)

            patients = data.get("data", [])
            for p in patients:
                patient_id = p.get("id")
                if patient_id is None:
                    yield p
                else:
                    yield self._get(f"/v1/patients/{patient_id}")

            cursor = data.get("nextPageCursor")
            if not cursor:
                break

    def iter_accounts(self, params: Optional[dict[str, Any]] = None):
        cursor = None

        while True:
            request_params: dict[str, Any] = {"limit": 100}
            if params:
                request_params.update(params)
            if cursor:
                request_params["pageCursor"] = cursor

            data = self._get("/v1/accounts", request_params)
            accounts = _extract_collection(data, ("accounts", "data", "items", "results"))
            for account in accounts:
                yield account

            cursor = data.get("nextPageCursor") if isinstance(data, dict) else None
            if not cursor:
                break

    def iter_patients_for_account(self, account_id: str):
        cursor = None

        while True:
            params: dict[str, Any] = {"limit": 100, "accountId": account_id}
            if cursor:
                params["pageCursor"] = cursor

            data = self._get("/v1/patients", params)
            patients = _extract_collection(data, ("patients", "data", "items", "results"))
            for patient in patients:
                yield patient

            cursor = data.get("nextPageCursor") if isinstance(data, dict) else None
            if not cursor:
                break

    def _find_accounts_by_client_code(self, client_code: str) -> list[dict[str, Any]]:
        matches: list[dict[str, Any]] = []
        seen_ids: set[str] = set()

        for field_name in ("pimsCode", "pimsId"):
            for account in self.iter_accounts({field_name: client_code}):
                account_id = str(account.get("id") or "")
                if account_id and account_id not in seen_ids:
                    matches.append(account)
                    seen_ids.add(account_id)

        return matches

    def _find_accounts_by_owner(self, owner_name: str, phone_no: str) -> list[dict[str, Any]]:
        candidates = list(self.iter_accounts({"name": owner_name}))
        owner_name_normalized = _normalize_lookup_text(owner_name)
        phone_digits = _normalize_phone_no(phone_no)

        exact_name_matches = [
            account
            for account in candidates
            if _normalize_lookup_text(_account_display_name(account)) == owner_name_normalized
        ]
        narrowed = exact_name_matches or candidates

        if not phone_digits:
            return narrowed

        phone_matches = []
        for account in narrowed:
            primary = account.get("primaryContact") or {}
            account_phone = _normalize_phone_no(_extract_phone_no((primary.get("communicationDetails") or [])))
            if account_phone and account_phone == phone_digits:
                phone_matches.append(account)

        return phone_matches or narrowed

    def find_patient(self, source_patient: ParsedPatientGroup) -> dict[str, Any]:
        accounts: list[dict[str, Any]] = []

        if source_patient.client:
            accounts = self._find_accounts_by_client_code(source_patient.client)

        if not accounts and source_patient.client_name:
            accounts = self._find_accounts_by_owner(source_patient.client_name, source_patient.phone_no)

        if not accounts:
            raise RuntimeError(
                f"Could not find Instinct account for client={source_patient.client!r} owner={source_patient.client_name!r}"
            )

        if len(accounts) > 1:
            account_ids = [account.get("id") for account in accounts]
            raise RuntimeError(
                f"Account lookup was ambiguous for client={source_patient.client!r} owner={source_patient.client_name!r}: {account_ids}"
            )

        account_id = accounts[0].get("id")
        if not account_id:
            raise RuntimeError(f"Matched account is missing id for client={source_patient.client!r}")

        wanted_name = _normalize_lookup_text(source_patient.patient_name)
        patient_matches = [
            patient
            for patient in self.iter_patients_for_account(str(account_id))
            if _normalize_lookup_text(patient.get("name")) == wanted_name
        ]

        if not patient_matches:
            raise RuntimeError(
                f"Could not find patient name={source_patient.patient_name!r} in account_id={account_id!r}"
            )

        if len(patient_matches) > 1:
            patient_ids = [patient.get("id") for patient in patient_matches]
            raise RuntimeError(
                f"Patient lookup was ambiguous for name={source_patient.patient_name!r} in account_id={account_id!r}: {patient_ids}"
            )

        patient_id = patient_matches[0].get("id")
        if patient_id is None:
            raise RuntimeError(
                f"Matched patient is missing id for name={source_patient.patient_name!r} in account_id={account_id!r}"
            )

        return self._get(f"/v1/patients/{patient_id}")

    def summarize_patient(self, patient_record):
        account = patient_record.get("account") or {}
        primary = account.get("primaryContact") or {}

        first = (primary.get("nameFirst") or "").strip()
        middle = (primary.get("nameMiddle") or "").strip()
        last = (primary.get("nameLast") or "").strip()

        name_parts = [part for part in (first, middle, last) if part]
        client_name = " ".join(name_parts) if name_parts else None

        phone_no = _extract_phone_no(primary.get("communicationDetails") or [])
        reminder_count = self.get_reminder_count_for_patient(patient_record.get("id"))

        return {
            "client_name": client_name,
            "patient_name": patient_record.get("name"),
            "phone_no": phone_no,
            "reminder_count": reminder_count,
            "patient_id": patient_record.get("id"),
            "account_id": patient_record.get("accountId"),
            "pims_code": patient_record.get("pimsCode"),
        }

# ---------------------------------------------------------------------------
# Audit helpers
# ---------------------------------------------------------------------------

def audit_all_patients(adapter: InstinctApiAdapter, limit: Optional[int] = None) -> list[dict[str, Any]]:
    adapter.token = adapter.authenticate()
    audit_rows: list[dict[str, Any]] = []

    for index, patient_record in enumerate(adapter.iter_patients(), start=1):
        if limit is not None and index > limit:
            break
        audit_rows.append(adapter.summarize_patient(patient_record))

    return audit_rows


# ---------------------------------------------------------------------------
# Output helpers
# ---------------------------------------------------------------------------


def plan_to_dict(plan: PatientImportPlan) -> dict[str, Any]:
    return {
        "patient": {
            "client": plan.patient.client,
            "client_name": plan.patient.client_name,
            "phone_no": plan.patient.phone_no,
            "patient_name": plan.patient.patient_name,
            "species": plan.patient.species,
            "breed": plan.patient.breed,
            "header_row_number": plan.patient.header_row_number,
        },
        "valid_reminders": [dataclasses.asdict(r) for r in plan.valid_reminders],
        "skipped_rows": [dataclasses.asdict(r) for r in plan.skipped_rows],
    }



def print_dry_run(plans: list[PatientImportPlan], max_patients: Optional[int] = None) -> None:
    shown = 0
    for plan in plans:
        if max_patients is not None and shown >= max_patients:
            break
        shown += 1

        patient = plan.patient
        print(f"Patient: {patient.patient_name} | Client: {patient.client} | Owner: {patient.client_name}")
        print(f"  Species/Breed: {patient.species} / {patient.breed}")
        print(f"  Source header row: {patient.header_row_number}")
        print("  Reminders to create:")
        if plan.valid_reminders:
            for reminder in plan.valid_reminders:
                print(
                    f"    - [{reminder.source_code}] {reminder.source_label} -> "
                    f"{reminder.mapped_label} @ {reminder.due_date} (row {reminder.row_number})"
                )
        else:
            print("    - none")

        print("  Skipped rows:")
        if plan.skipped_rows:
            for skipped in plan.skipped_rows:
                print(
                    f"    - [{skipped.source_code}] {skipped.source_label} | "
                    f"{skipped.reason} (row {skipped.row_number})"
                )
        else:
            print("    - none")
        print()


# ---------------------------------------------------------------------------
# Import execution
# ---------------------------------------------------------------------------


def execute_import(
    plans: list[PatientImportPlan],
    adapter: InstinctApiAdapter,
    max_patients: Optional[int] = None,
) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    adapter.token = adapter.authenticate()

    processed = 0
    for plan in plans:
        if max_patients is not None and processed >= max_patients:
            break
        processed += 1

        if not plan.valid_reminders:
            LOGGER.info(
                "Skipping patient %s because there are no valid reminders after mapping/filtering",
                plan.patient.patient_name,
            )
            results.append(
                {
                    "patient_name": plan.patient.patient_name,
                    "client": plan.patient.client,
                    "status": "skipped_no_valid_reminders",
                }
            )
            continue

        patient_record = adapter.find_patient(plan.patient)
        response = adapter.add_reminders(patient_record, plan.valid_reminders)
        results.append(
            {
                "patient_name": plan.patient.patient_name,
                "client": plan.patient.client,
                "status": "submitted",
                "response": response,
            }
        )

    return results


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Parse EVH reminder spreadsheet and prepare Instinct reminder import")
    parser.add_argument("spreadsheet", type=Path, nargs="?", help="Path to the .xlsx reminder file")
    parser.add_argument("--sheet", type=str, default=None, help="Optional worksheet name")
    parser.add_argument("--dry-run", action="store_true", help="Print parsed/import plans and do not call the API")
    parser.add_argument("--export-json", type=Path, default=None, help="Write parsed plans or audit rows to JSON")
    parser.add_argument("--max-patients", type=int, default=None, help="Limit processing to first N patient groups or audit rows")
    parser.add_argument("--execute", action="store_true", help="Execute import via Instinct API adapter")
    parser.add_argument("--audit-patients", action="store_true", help="Enumerate live patients and output client name, patient name, phone number, and current reminder count")
    parser.add_argument("--base-url", type=str, default=os.getenv("INSTINCT_API_BASE_URL", ""))
    parser.add_argument("--username", type=str, default=os.getenv("INSTINCT_API_USERNAME", ""))
    parser.add_argument("--password", type=str, default=os.getenv("INSTINCT_API_PASSWORD", ""))
    parser.add_argument("--log-level", type=str, default="INFO")
    return parser



def main() -> int:
    parser = build_arg_parser()
    args = parser.parse_args()

    logging.basicConfig(
        level=getattr(logging, args.log_level.upper(), logging.INFO),
        format="%(levelname)s %(name)s: %(message)s",
    )

    if args.audit_patients:
        missing = [name for name, value in {
            "base_url": args.base_url,
            "username": args.username,
            "password": args.password,
        }.items() if not value]
        if missing:
            parser.error(f"Missing required API configuration for audit mode: {', '.join(missing)}")

        adapter = InstinctApiAdapter(
            base_url=args.base_url,
            username=args.username,
            password=args.password,
        )
        audit_rows = audit_all_patients(adapter, limit=args.max_patients)

        if args.export_json:
            args.export_json.write_text(json.dumps(audit_rows, indent=2), encoding="utf-8")
            LOGGER.info("Wrote audit JSON to %s", args.export_json)

        print(json.dumps(audit_rows, indent=2))
        return 0

    if not args.spreadsheet:
        parser.error("spreadsheet is required unless --audit-patients is used")

    groups = parse_grouped_spreadsheet(args.spreadsheet, worksheet_name=args.sheet)
    plans = build_import_plan(groups)

    if args.export_json:
        payload = [plan_to_dict(plan) for plan in plans[: args.max_patients]] if args.max_patients else [plan_to_dict(plan) for plan in plans]
        args.export_json.write_text(json.dumps(payload, indent=2), encoding="utf-8")
        LOGGER.info("Wrote JSON plan export to %s", args.export_json)

    if args.dry_run or not args.execute:
        print_dry_run(plans, max_patients=args.max_patients)
        return 0

    missing = [name for name, value in {
        "base_url": args.base_url,
        "username": args.username,
        "password": args.password,
    }.items() if not value]
    if missing:
        parser.error(f"Missing required API configuration for execute mode: {', '.join(missing)}")

    adapter = InstinctApiAdapter(
        base_url=args.base_url,
        username=args.username,
        password=args.password,
    )
    results = execute_import(plans, adapter, max_patients=args.max_patients)
    print(json.dumps(results, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
