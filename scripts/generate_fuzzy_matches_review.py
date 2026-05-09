import csv
import difflib
import re

from openpyxl import load_workbook


XLSX = "EVHInventorySuppliers.xlsx"
STOCK = "Stockroom · Instinct Stockroom.csv"
EXACT = "Stockroom · Instinct Stockroom_with_Avimark_EVH_Suppliers_exact.csv"
OUT = "docs/evh_fuzzy_matches_review.csv"


def norm(value: str) -> str:
    value = (value or "").lower().strip()
    value = value.lstrip("📦 ")
    value = value.replace("&", " and ")
    return re.sub(r"[^a-z0-9]+", "", value)


def main() -> None:
    with open(STOCK, newline="", encoding="utf-8-sig") as f:
        stock_rows = list(csv.DictReader(f))

    with open(EXACT, newline="", encoding="utf-8-sig") as f:
        exact_rows = {row["PIMS ID"]: row for row in csv.DictReader(f)}

    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb.active
    wb_rows = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx == 1:
            continue
        code = "" if len(row) < 1 or row[0] is None else str(row[0]).strip()
        desc = "" if len(row) < 2 or row[1] is None else str(row[1]).strip()
        supplier = "" if len(row) < 3 or row[2] is None else str(row[2]).strip()
        if code or desc:
            wb_rows.append((code, desc, supplier))

    rows = []
    for srow in stock_rows:
        sid = srow["PIMS ID"].strip()
        if (exact_rows.get(sid, {}).get("Suppliers") or "").strip():
            continue

        sn = norm(srow["Product"])
        best = None
        best_score = 0.0
        for code, desc, supplier in wb_rows:
            score = difflib.SequenceMatcher(None, sn, norm(desc)).ratio()
            if score > best_score:
                best_score = score
                best = (code, desc, supplier)
        if best and best_score >= 0.84:
            code, desc, supplier = best
            rows.append(
                {
                    "XLS ID": code,
                    "XLS Product": desc,
                    "XLS Supplier": supplier,
                    "Stockroom ID": sid,
                    "Stockroom Product": srow["Product"].strip(),
                    "Supplier To Use": supplier,
                    "Score": f"{best_score:.3f}",
                }
            )

    with open(OUT, "w", newline="", encoding="utf-8") as f:
        fieldnames = [
            "XLS ID",
            "XLS Product",
            "XLS Supplier",
            "Stockroom ID",
            "Stockroom Product",
            "Supplier To Use",
            "Score",
        ]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(OUT)
    print(f"rows={len(rows)}")


if __name__ == "__main__":
    main()
