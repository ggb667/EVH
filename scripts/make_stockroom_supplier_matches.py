import csv
import re
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

SRC = Path("Instinct_Stockroom.csv")
XLSX = Path("EVHInventorySuppliers.xlsx")
OUT = Path("Instinct_Stockroom_with_supplier_matches.csv")
SUPPLIERS_CSV = Path("stockroom_suppliers_ids.csv")
EXPECTED_COLUMNS = 18
DESCRIPTION_COLUMN = "Product"
SUPPLIER_COLUMN = "Suppliers"
ID_COLUMNS = ("PIMS ID", "Manufacturer")


def norm_text(value: str) -> str:
    value = (value or "").strip().lower()
    value = value.lstrip("📦 ")
    value = value.replace("&", " and ")
    return re.sub(r"[^a-z0-9]+", "", value)


def normalize_id(value: str) -> str:
    return re.sub(r"\s+", "", (value or "").strip())


def normalize_supplier_label(value: str) -> str:
    value = (value or "").strip().lower()
    value = value.replace("&", " and ")
    value = value.replace("/", " ")
    value = re.sub(r"[\(\)\[\],.]", " ", value)
    value = re.sub(r"\bsupp\b", "supply", value)
    value = re.sub(r"\bvets\b", "veterinary", value)
    value = re.sub(r"\bvet\b", "veterinary", value)
    value = re.sub(r"\binc\b", "", value)
    value = re.sub(r"\bincorporated\b", "", value)
    value = re.sub(r"\bllc\b", "", value)
    value = re.sub(r"\bltd\b", "", value)
    value = re.sub(r"\bco\b", "", value)
    value = re.sub(r"\bcorp\b", "", value)
    value = re.sub(r"\bcompany\b", "", value)
    value = re.sub(r"\bservices\b", "service", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def split_ids(value: str) -> list[str]:
    if not value:
        return []
    pieces = re.split(r"[\/,;|]", value)
    return [normalize_id(piece) for piece in pieces if normalize_id(piece)]


def ensure_expected_width(row: list[str], *, source: Path, line_no: int) -> None:
    if len(row) != EXPECTED_COLUMNS:
        raise ValueError(
            f"{source} line {line_no} has {len(row)} columns; expected {EXPECTED_COLUMNS}"
        )


@dataclass
class WorkbookMatch:
    suppliers: set[str] = field(default_factory=set)
    ids: set[str] = field(default_factory=set)
    descriptions: set[str] = field(default_factory=set)


def load_supplier_canonical_map(path: Path) -> dict[str, str]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        canonical: dict[str, str] = {}
        for row in reader:
            label = (row.get("label") or "").strip()
            if not label:
                continue
            key = normalize_supplier_label(label)
            canonical.setdefault(key, label)
    return canonical


def load_stockroom_rows(path: Path) -> tuple[list[str], list[dict[str, str]], dict[str, list[int]]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if header is None:
            raise ValueError(f"{path} is empty")
        ensure_expected_width(header, source=path, line_no=1)

        rows: list[dict[str, str]] = []
        descriptions_seen: dict[str, list[int]] = defaultdict(list)
        for line_no, row in enumerate(reader, start=2):
            ensure_expected_width(row, source=path, line_no=line_no)
            mapped = dict(zip(header, row))
            desc = mapped[DESCRIPTION_COLUMN].strip()
            if desc:
                descriptions_seen[desc].append(line_no)
            rows.append(mapped)
    duplicates = {desc: lines for desc, lines in descriptions_seen.items() if len(lines) > 1}
    return header, rows, duplicates


def load_workbook_matches(path: Path) -> dict[str, WorkbookMatch]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:  # pragma: no cover - environment guard
        raise ModuleNotFoundError(
            "openpyxl is required to read EVHInventorySuppliers.xlsx; install it in the project environment"
        ) from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    matches: dict[str, WorkbookMatch] = defaultdict(WorkbookMatch)
    for line_no, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if line_no == 1:
            continue

        code = normalize_id("" if len(row) < 1 or row[0] is None else str(row[0]))
        desc = "" if len(row) < 2 or row[1] is None else str(row[1]).strip()
        supplier = "" if len(row) < 3 or row[2] is None else str(row[2]).strip()
        other_ids = split_ids("" if len(row) < 4 or row[3] is None else str(row[3]))

        if not code and not desc and not supplier and not other_ids:
            continue

        norm_desc = norm_text(desc)
        if norm_desc:
            matches[norm_desc].descriptions.add(desc)
        if supplier:
            matches[norm_desc].suppliers.add(supplier)
        if code:
            matches[code].ids.add(code)
            if supplier:
                matches[code].suppliers.add(supplier)
            if desc:
                matches[code].descriptions.add(desc)
        for item_id in other_ids:
            if item_id:
                matches[item_id].ids.add(item_id)
                if supplier:
                    matches[item_id].suppliers.add(supplier)
                if desc:
                    matches[item_id].descriptions.add(desc)

    return matches


def canonicalize_suppliers(suppliers: set[str], canonical_map: dict[str, str]) -> list[str]:
    resolved: list[str] = []
    seen = set()
    for supplier in sorted(suppliers):
        key = normalize_supplier_label(supplier)
        canonical = canonical_map.get(key)
        if canonical is None:
            # fall back to the closest canonical label when the supplier text is abbreviated
            candidates = sorted(
                canonical_map.items(),
                key=lambda item: (1 - len(set(key) & set(item[0])) / max(len(set(key) | set(item[0])), 1), len(item[1]), item[1]),
            )
            canonical = candidates[0][1] if candidates else supplier
        if canonical not in seen:
            seen.add(canonical)
            resolved.append(canonical)
    return resolved


def collect_suppliers(row: dict[str, str], workbook_matches: dict[str, WorkbookMatch]) -> list[str]:
    suppliers: set[str] = set()
    candidate_keys = {
        normalize_id(row.get("PIMS ID", "")),
        normalize_id(row.get("Manufacturer", "")),
        norm_text(row.get(DESCRIPTION_COLUMN, "")),
    }

    for key in candidate_keys:
        if not key:
            continue
        match = workbook_matches.get(key)
        if match:
            suppliers.update(match.suppliers)

    return sorted(suppliers)


def write_rows(
    path: Path,
    header: list[str],
    rows: list[dict[str, str]],
    workbook_matches: dict[str, WorkbookMatch],
    canonical_map: dict[str, str],
) -> int:
    matched_rows: list[list[str]] = []
    unmatched_rows: list[list[str]] = []

    for row in rows:
        suppliers = canonicalize_suppliers(collect_suppliers(row, workbook_matches), canonical_map)
        out_row = [row.get(column, "") for column in header]
        out_row[header.index(SUPPLIER_COLUMN)] = ", ".join(suppliers)

        desc_key = norm_text(row.get(DESCRIPTION_COLUMN, ""))
        id_key = normalize_id(row.get("PIMS ID", ""))
        mfg_key = normalize_id(row.get("Manufacturer", ""))
        is_match = bool(
            suppliers
            or (desc_key and desc_key in workbook_matches)
            or (id_key and id_key in workbook_matches)
            or (mfg_key and mfg_key in workbook_matches)
        )

        if is_match:
            matched_rows.append(out_row)
        else:
            unmatched_rows.append(out_row)

    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for row in matched_rows:
            ensure_expected_width(row, source=path, line_no=0)
            writer.writerow(row)
        f.write("\n")
        for row in unmatched_rows:
            ensure_expected_width(row, source=path, line_no=0)
            writer.writerow(row)

    return len(matched_rows)


def main() -> None:
    header, rows, duplicates = load_stockroom_rows(SRC)
    workbook_matches = load_workbook_matches(XLSX)
    canonical_map = load_supplier_canonical_map(SUPPLIERS_CSV)
    matched_count = write_rows(OUT, header, rows, workbook_matches, canonical_map)
    print(OUT)
    print(f"matched_rows={matched_count}")
    print(f"unmatched_rows={len(rows) - matched_count}")
    if duplicates:
        print(f"duplicate_descriptions={len(duplicates)}")


if __name__ == "__main__":
    main()
